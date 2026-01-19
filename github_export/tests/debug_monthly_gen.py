import json
import os
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage
from datetime import datetime

# Paths
BASE_DIR = r"C:\Users\raven.yeh\Desktop\專案拓展部\09_AI增能\行政助手(vscode重建)\work_assistant"
PROJECT_ID = "5f078857-f623-4612-a601-99105a83e338"
CONFIG_PATH = os.path.join(BASE_DIR, "projects", PROJECT_ID, "config.json")
ENTRIES_PATH = os.path.join(BASE_DIR, "projects", PROJECT_ID, "entries.json")
UPLOADS_DIR = os.path.join(BASE_DIR, "uploads")

print(f"Loading config from {CONFIG_PATH}")
with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
    config = json.load(f)

print(f"Loading entries from {ENTRIES_PATH}")
with open(ENTRIES_PATH, 'r', encoding='utf-8') as f:
    entries = json.load(f)

template_file = config['template_file']
template_path = os.path.join(UPLOADS_DIR, template_file)

print(f"Loading template from {template_path}")
if not os.path.exists(template_path):
    print("Template not found!")
    exit(1)

wb = openpyxl.load_workbook(template_path)
source_sheet = wb.worksheets[0]

# Sort entries by date
entries.sort(key=lambda x: x['date'])

print(f"Found {len(entries)} entries")

for entry in entries:
    print(f"Processing entry: {entry['date']}")
    target_sheet = wb.copy_worksheet(source_sheet)
    
    # 1. Determine Sheet Name
    sheet_name = entry['date'][5:] 
    
    # Try to find specific sheet_name param
    for param in config.get('parameters', []):
            if param['name'] == 'sheet_name' and entry['data'].get('sheet_name'):
                raw_name = entry['data']['sheet_name']
                if raw_name: sheet_name = raw_name
                break
    
    # Sanitize sheet name
    safe_name = str(sheet_name).replace(':', '').replace('/', '-').replace('\\', '').replace('?', '').replace('*', '').replace('[', '').replace(']', '')[:30]
    target_sheet.title = safe_name
    print(f"  > Created sheet: {safe_name}")
    
    # 2. Fill Data & Images
    data_map = entry['data']
    
    for param in config.get('parameters', []):
        key = param['name']
        val = data_map.get(key)
        
        if not val: continue
        
        if param['type'] == 'image':
            # Handle Image
            anchor = param.get('style', {}).get('anchor_cell')
            if anchor and val:
                try:
                    # Resolve path
                    # val is like "uploads/filename.jpg"
                    # But in the entry data it might be relative to project root or absolute?
                    # Previous code: entry_data['data'][field] = f"uploads/{rel_path}"
                    
                    # If val starts with "uploads/", we need to adjust
                    if val.startswith("uploads/"):
                         # The val is "uploads/filename"
                         # UPLOADS_DIR is ".../work_assistant/uploads"
                         # So we need to join BASE_DIR (work_assistant) with val?
                         # app.config['UPLOAD_FOLDER'] is usually ".../work_assistant/uploads"
                         # And rel_path was rel to that.
                         
                         # If val is "uploads/image.jpg", and we are in work_assistant, 
                         # then os.path.join(BASE_DIR, val) should be correct if BASE_DIR is work_assistant root?
                         # Wait, uploads is inside work_assistant.
                         img_path = os.path.join(BASE_DIR, val) # work_assistant/uploads/...
                    else:
                         img_path = val
                    
                    print(f"    > Inserting image: {img_path} at {anchor}")
                    
                    if os.path.exists(img_path):
                        img = OpenpyxlImage(img_path)
                        
                        if ',' in str(anchor):
                            r, c = map(int, anchor.split(','))
                            cell_addr = f"{get_column_letter(c)}{r}"
                        else:
                            cell_addr = "A1"
                        
                        img.anchor = cell_addr
                        target_sheet.add_image(img)
                        
                        # Clear text
                        if ',' in str(anchor):
                            r, c = map(int, anchor.split(','))
                            target_sheet.cell(row=r, column=c).value = ""
                    else:
                        print(f"    ! Image file missing: {img_path}")
                except Exception as img_err:
                    print(f"    ! Image Error: {img_err}")

        else:
            # Handle Text Replacement
            tag = f"{{{{ {key} }}}}"
            
            # Search and replace
            for row in target_sheet.iter_rows():
                for cell in row:
                        if cell.value and isinstance(cell.value, str) and tag in cell.value:
                            target_sheet.cell(row=cell.row, column=cell.column).value = str(cell.value).replace(tag, str(val))

# Remove original
if len(wb.sheetnames) > 1:
    wb.remove(source_sheet)

out_name = "Debug_Report.xlsx"
out_path = os.path.join(UPLOADS_DIR, out_name)
wb.save(out_path)
print(f"Saved to {out_path}")
