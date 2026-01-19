import openpyxl
import os

# Path to the generated reproduction file
repro_file = r"C:\Users\raven.yeh\Desktop\專案拓展部\09_AI增能\行政助手(vscode重建)\work_assistant\uploads\Reproduction_Report_v2.xlsx"

if not os.path.exists(repro_file):
    print(f"File not found: {repro_file}")
else:
    print(f"Checking file: {repro_file}")
    wb = openpyxl.load_workbook(repro_file)
    # The reproduction script named the sheet "Test_Photos"
    if "Test_Photos" in wb.sheetnames:
        ws = wb["Test_Photos"]
        print(f"Sheet: {ws.title}")
        # accessing protected member _images for debugging purposes
        if hasattr(ws, '_images'):
            print(f"Number of images detected (via _images): {len(ws._images)}")
            for i, img in enumerate(ws._images):
                # img.anchor can be a string or object depending on openpyxl version
                print(f"  Image {i+1}: Anchor={img.anchor}") 
                print(f"  Dimensions: {img.width}x{img.height}")
        else:
            print("Attribute _images not found on worksheet.")
    else:
        print("Sheet 'Test_Photos' not found.")
        print("Available sheets:", wb.sheetnames)
        
    wb.close()
