import openpyxl
import os

# Paths
BASE_DIR = r"C:\Users\raven.yeh\Desktop\專案拓展部\09_AI增能\行政助手(vscode重建)\work_assistant"
UPLOADS_DIR = os.path.join(BASE_DIR, "uploads")
config_path = os.path.join(BASE_DIR, "projects", "5f078857-f623-4612-a601-99105a83e338", "config.json")

import json
with open(config_path, 'r', encoding='utf-8') as f:
    config = json.load(f)

template_file = config['template_file']
template_path = os.path.join(UPLOADS_DIR, template_file)

print(f"Checking layout of {template_path}")
try:
    wb = openpyxl.load_workbook(template_path)
    ws = wb.worksheets[0]

    # Helper to get range dimensions
    def get_range_dimensions(sheet, range_string):
        min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(range_string)
        
        total_width = 0
        for col_idx in range(min_col, max_col + 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            col_w = sheet.column_dimensions[col_letter].width
            # If width is missing, openpyxl returns None. Excel default is ~8.43
            if col_w is None: col_w = 8.43
            total_width += col_w
            
        total_height = 0
        for row_idx in range(min_row, max_row + 1):
            row_h = sheet.row_dimensions[row_idx].height
            # If height is missing, default is ~15
            if row_h is None: row_h = 15
            total_height += row_h
            
        # Refined Approximations
        # Width: 1 unit approx 7-7.5 pixels depending on font. Let's use 7.2
        # Height: 1 point = 1.333 pixels
        
        width_px = total_width * 7.5 # Adjusted slightly up for margin safety
        height_px = total_height * 1.333
        
        return width_px, height_px

    print("Merged Cells:")
    target_ranges = {}
    
    # Check specifically for A2 and F2 (row 2, col 1 and row 2, col 6)
    # Note: openpyxl merged_cells is a list of CellRange
    
    found_A2 = False
    found_F2 = False
    
    for rng in ws.merged_cells.ranges:
        # Check A2 (1,2)
        if rng.min_col <= 1 and rng.max_col >= 1 and rng.min_row <= 2 and rng.max_row >= 2:
            target_ranges['A2'] = str(rng)
            found_A2 = True
        
        # Check F2 (6,2)
        if rng.min_col <= 6 and rng.max_col >= 6 and rng.min_row <= 2 and rng.max_row >= 2:
            target_ranges['F2'] = str(rng)
            found_F2 = True
            
    if not found_A2: target_ranges['A2'] = "A2"
    if not found_F2: target_ranges['F2'] = "F2"

    for key, rng_str in target_ranges.items():
        w, h = get_range_dimensions(ws, rng_str)
        print(f"Target {key} ({rng_str}): Approx {w:.1f} x {h:.1f} pixels")
except Exception as e:
    print(f"Error: {e}")
