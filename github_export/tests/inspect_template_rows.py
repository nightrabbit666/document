import openpyxl
import os

BASE_DIR = r"C:\Users\raven.yeh\Desktop\專案拓展部\09_AI增能\行政助手(vscode重建)\work_assistant"
UPLOADS_DIR = os.path.join(BASE_DIR, "uploads")
config_path = os.path.join(BASE_DIR, "projects", "5f078857-f623-4612-a601-99105a83e338", "config.json")

import json
with open(config_path, 'r', encoding='utf-8') as f:
    config = json.load(f)

template_file = config['template_file']
template_path = os.path.join(UPLOADS_DIR, template_file)

wb = openpyxl.load_workbook(template_path)
ws = wb.worksheets[0]

print("--- Cell Values (A1:A10) ---")
for i in range(1, 11):
    cell = ws.cell(row=i, column=1)
    print(f"A{i}: {repr(cell.value)}")

print("\n--- Cell Values (F1:F10) ---")
for i in range(1, 11):
    cell = ws.cell(row=i, column=6)
    print(f"F{i}: {repr(cell.value)}")

print("\n--- Merged Cells containing A data ---")
for rng in ws.merged_cells.ranges:
    if rng.min_col <= 1 and rng.max_col >= 1:
        print(rng)
        
print("\n--- Row Heights (1-20) ---")
for i in range(1, 21):
    val = ws.row_dimensions[i].height
    print(f"Row {i}: {val}")
