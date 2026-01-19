import json
import os
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage
import shutil

# Paths
BASE_DIR = r"C:\Users\raven.yeh\Desktop\專案拓展部\09_AI增能\行政助手(vscode重建)\work_assistant"
TEST_PHOTOS_DIR = r"C:\Users\raven.yeh\Desktop\專案拓展部\09_AI增能\行政助手(vscode重建)\參考\照片"
PROJECT_ID = "5f078857-f623-4612-a601-99105a83e338"
CONFIG_PATH = os.path.join(BASE_DIR, "projects", PROJECT_ID, "config.json")
UPLOADS_DIR = os.path.join(BASE_DIR, "uploads")
OUTPUT_PATH = os.path.join(UPLOADS_DIR, "Reproduction_Report_v2.xlsx")

# 1. Load Config
print(f"Loading config from {CONFIG_PATH}")
with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
    config = json.load(f)

# 2. Mock Entry Data
entry_data = {
    "date": "2026-02-01",
    "data": {
        "sheet_name": "Test Photos",
        "removal_photo": os.path.join(TEST_PHOTOS_DIR, "1225包裝材ˊ清運照片.png"),
        "manifest_photo": os.path.join(TEST_PHOTOS_DIR, "1225方形單.png")
    }
}

print("Mock Data:", entry_data)

# 3. Load Template
template_file = config['template_file']
template_path = os.path.join(UPLOADS_DIR, template_file)
print(f"Loading template from {template_path}")

try:
    wb = openpyxl.load_workbook(template_path)
    source_sheet = wb.worksheets[0]
    target_sheet = wb.copy_worksheet(source_sheet)
    target_sheet.title = "Test_Photos"
    
    # 4. Fill Logic (Copied from txtapp.py)
    data_map = entry_data['data']
    
    for param in config.get('parameters', []):
        key = param['name']
        val = data_map.get(key)
        original_text = param.get('original_text', '')
        tag = f"{{{{ {key} }}}}"
        
        if param['type'] == 'image':
            anchor = param.get('style', {}).get('anchor_cell')
            image_inserted = False
            
            if val:
                try:
                    # Direct file path usage for this test
                    img_path = val 
                    print(f"Checking image: {img_path}")
                    
                    if os.path.exists(img_path):
                        img = OpenpyxlImage(img_path)
                        
                        # Auto-Resize Logic (New addition to fix ratio issues?)
                        print(f"  > Original size: {img.width}x{img.height}")
                        
                        if anchor:
                            if ',' in str(anchor):
                                r, c = map(int, anchor.split(','))
                                cell_addr = f"{get_column_letter(c)}{r}"
                            else:
                                cell_addr = "A1"
                        else:
                            cell_addr = "A1"
                        
                        img.anchor = cell_addr
                        target_sheet.add_image(img)
                        image_inserted = True
                        print(f"  > Inserted at {cell_addr}")
                    else:
                        print(f"  > File not found: {img_path}")
                except Exception as e:
                    print(f"  > Error inserting image: {e}")
    
            # Cleanup text
            replacement_text = "" if image_inserted else original_text
            for row in target_sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and tag in cell.value:
                        if cell.value.strip() == tag:
                            cell.value = replacement_text
                        else:
                            cell.value = cell.value.replace(tag, replacement_text)
    
       
    # Remove template
    if len(wb.sheetnames) > 1:
        wb.remove(source_sheet)
    
    wb.save(OUTPUT_PATH)
    print(f"Saved to {OUTPUT_PATH}")
except Exception as e:
    import traceback
    traceback.print_exc()
