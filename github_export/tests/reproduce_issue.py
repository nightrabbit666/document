import os
import json
import logging
import google.generativeai as genai
import openpyxl
from deepdiff import DeepDiff
from dotenv import load_dotenv

# Setup basic logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Load env (for API Key)
load_dotenv()
api_key = os.getenv('GEMINI_API_KEY')
if not api_key:
    print("ERROR: No API Key found.")
    exit(1)

genai.configure(api_key=api_key)

# Define Paths
# Fix path: .env is in parent directory relative to tests/
load_dotenv(os.path.join(os.path.dirname(os.path.dirname(__file__)), '.env'))

# Fix Paths: Reference folder is in parent directory
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) # Go up one level from tests/
REF_DIR = os.path.join(BASE_DIR, "參考")
TEMPLATE_PATH = os.path.join(REF_DIR, "空白.xlsx")
FILLED_PATH = os.path.join(REF_DIR, "台中美光一廠出貨紀錄表-2025年12月(1201-1215).xlsx")

print(f"Testing with:\nTemplate: {TEMPLATE_PATH}\nFilled: {FILLED_PATH}")

def extract_xlsx_structure(filepath, limit_rows=100):
    """提取 Excel 文件的結構化資訊 (Sheet 與 Cell) [Copied from txtapp.py]"""
    if not filepath or not os.path.exists(filepath):
        print(f"File not found: {filepath}")
        return {}

    try:
        # Load without read_only to access images. data_only=True for Formula values.
        wb = openpyxl.load_workbook(filepath, data_only=True)
        # Add metadata for analysis
        structure = {
            "sheet_names": wb.sheetnames,
            "sheets": {}
        }
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            cells_data = {}
            row_count = 0
            
            # 1. Extract Text Data
            for row in ws.iter_rows():
                if row_count > limit_rows: break
                for cell in row:
                    if cell.value is not None:
                        # 記錄坐標與值
                        cells_data[f"{cell.row},{cell.column}"] = str(cell.value).strip()
                row_count += 1
            
            # 2. Extract Images (Marker only)
            try:
                # openpyxl 3.0+ uses ws._images or ws.images
                images_list = getattr(ws, '_images', []) or getattr(ws, 'images', [])
                print(f"Sheet '{sheet_name}' has {len(images_list)} images.")
                
                for img in images_list:
                    # Attempt to find anchor (Top-Left)
                    r, c = None, None
                    
                    # Handling different anchor types (OneCell, TwoCell, Absolute)
                    anchor = img.anchor
                    if hasattr(anchor, '_from'): # TwoCellAnchor
                        r = anchor._from.row + 1 # 0-index to 1-index
                        c = anchor._from.col + 1
                    elif hasattr(anchor, 'row'): # OneCellAnchor (sometimes)
                        r = anchor.row + 1
                        c = anchor.col + 1
                    
                    print(f"Image found at Row:{r}, Col:{c}")
                        
                    if r and c:
                        key = f"{r},{c}"
                        exist_val = cells_data.get(key, "")
                        
                        # Extract dimensions avoiding distoration
                        w = getattr(img, 'width', 0)
                        h = getattr(img, 'height', 0)
                        marker = f"<<IMAGE_PRESENT|W:{w}|H:{h}>>"
                        
                        cells_data[key] = f"{exist_val} {marker}".strip()
                        print(f"Injected marker at {key}: {cells_data[key]}")
                        
            except Exception as img_err:
                logger.warning(f"Excel Image extraction warning: {img_err}")
            
            structure["sheets"][sheet_name] = {
                "cells": cells_data
            }
        return structure
    except Exception as e:
        logger.error(f"Excel structure error: {e}")
        return {}

# 1. Extract
print("Extracting Template...")
blank_structure = extract_xlsx_structure(TEMPLATE_PATH)
print("Extracting Filled...")
filled_structure = extract_xlsx_structure(FILLED_PATH)

# 2. DeepDiff
print("Running DeepDiff...")
diff = DeepDiff(blank_structure, filled_structure, ignore_order=False, view='tree')
changes = []

if 'values_changed' in diff:
    for node in diff['values_changed']:
        path = " -> ".join([str(k) for k in node.path(output_format='list')])
        old_val = node.t1
        new_val = node.t2
        changes.append(f"[內容變更] 位置: {path} | 原始: '{old_val}' -> 填寫: '{new_val}'")

if 'dictionary_item_added' in diff:
    for node in diff['dictionary_item_added']:
        path = " -> ".join([str(k) for k in node.path(output_format='list')])
        val = node.t2
        changes.append(f"[新增內容] 位置: {path} | 內容: '{val}'")

# --- Deep Analysis of Sheet Logic ---
print("Analyzing Sheet Logic...")
template_sheets = blank_structure.get("sheet_names", [])
filled_sheets = filled_structure.get("sheet_names", [])

print(f"Template Sheets: {template_sheets}")
print(f"Filled Sheets: {filled_sheets}")

sheet_diff = [s for s in filled_sheets if s not in template_sheets]
print(f"New Sheets Detected: {sheet_diff}")

formatted_diff_report = ""
if changes:
    print(f"DeepDiff found {len(changes)} changes.")
    formatted_diff_report = "\n".join(changes)
else:
    print("DeepDiff found ZERO changes.")
    formatted_diff_report = "無顯著結構差異 (Fallback mode)"

# 3. AI Inference
print("Calling AI...")
template_type = "excel"
try:
    # TRYING DIFFERENT MODELS
    model_name = 'gemini-3-flash-preview' # Upgraded
    print(f"Using Model: {model_name}")
    
    model = genai.GenerativeModel(model_name)
    
    # Minimize JSON for prompt efficiency if too large
    blank_json = json.dumps(blank_structure, ensure_ascii=False)[:30000]
    filled_json = json.dumps(filled_structure, ensure_ascii=False)[:30000]

    prompt = f"""
    你是一個高階文檔自動化架構師。請針對這份「月報表」進行深度邏輯分析。
    
    【已知情境】
    - 這是一份每日填寫的月報表。
    - 使用者會依照日期新增工作表 (Sheet)。
    - 工作表名稱可能包含日期 (如 '1201', '1202')。
    - 每天的填寫項目可能不同。

    【輸入結構】
    1. 模板工作表清單: {template_sheets}
    2. 已填寫工作表清單: {filled_sheets}
    3. 新增的工作表: {sheet_diff}
    
    【分析任務】
    1. **識別製表邏輯**: 請說明這份報表是如何擴增的？(例如：「依照日期 MM/DD 建立新分頁」)。
    2. **識別日期模式**: 從工作表名稱中提取日期規則。
    3. **識別變數**: 即使 DeepDiff 沒抓到，請從 JSON 結構中找出「應填寫欄位」(尋找 '照片', '數量', '項目' 等關鍵字)。
    4. **計算 Token**: 請估算本次分析約使用了多少 Token (僅供參考)。

    請以 JSON 回傳:
    {{
       "logic_summary": "說明製表邏輯...",
       "date_pattern": "說明日期規則...",
       "parameters": [ {{ "name": "...", "description": "...", "type": "..." }} ],
       "token_usage_estimate": "..."
    }}
    """
    
    response = model.generate_content(prompt)
    print("AI Response Received:")
    print(response.text)
    print(f"Usage Metadata: {response.usage_metadata}")

except Exception as e:
    print(f"AI Error: {e}")
