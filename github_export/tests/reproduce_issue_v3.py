import json
import os
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage
import shutil

# Paths
BASE_DIR = r"C:\Users\raven.yeh\Desktop\專案拓展部\09_AI增能\行政助手(vscode重建)\work_assistant"
TEST_PHOTOS_DIR = r"C:\Users\raven.yeh\Desktop\專案拓展部\09_AI增能\行政助手(vscode重建)\參考\照片"
PROJECT_ID = "5f078857-f623-4612-a601-99105a83e338"
CONFIG_PATH = os.path.join(BASE_DIR, "projects", PROJECT_ID, "config.json")
UPLOADS_DIR = os.path.join(BASE_DIR, "uploads")
OUTPUT_PATH = os.path.join(UPLOADS_DIR, "Reproduction_Report_v3.xlsx")

# Ensure UPLOADs dir exists
if not os.path.exists(UPLOADS_DIR):
    os.makedirs(UPLOADS_DIR)

# 1. Copy Files with Simple Names
src_img1 = os.path.join(TEST_PHOTOS_DIR, "1225包裝材ˊ清運照片.png")
src_img2 = os.path.join(TEST_PHOTOS_DIR, "1225方形單.png")

dst_img1 = os.path.join(UPLOADS_DIR, "test_img1.png")
dst_img2 = os.path.join(UPLOADS_DIR, "test_img2.png")

shutil.copy2(src_img1, dst_img1)
shutil.copy2(src_img2, dst_img2)
print("Copied images to simple filenames.")

# 2. Load Config & Template
with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
    config = json.load(f)

template_file = config['template_file']
template_path = os.path.join(UPLOADS_DIR, template_file)
print(f"Loading template: {template_path}")
print(f"Template Size: {os.path.getsize(template_path)} bytes")

wb = openpyxl.load_workbook(template_path)
source_sheet = wb.worksheets[0]
target_sheet = wb.copy_worksheet(source_sheet)
target_sheet.title = "Test_Photos_V3"

# 3. Insert Images
# Simulate removal_photo -> A2
img1 = OpenpyxlImage(dst_img1)
img1.anchor = "A2"
target_sheet.add_image(img1)
print(f"Inserted img1 at A2 ({img1.width}x{img1.height})")

# Simulate manifest_photo -> F2
img2 = OpenpyxlImage(dst_img2)
img2.anchor = "F2"
target_sheet.add_image(img2)
print(f"Inserted img2 at F2 ({img2.width}x{img2.height})")

# Test: Clear text
for row in target_sheet.iter_rows():
    for cell in row:
        if cell.value and isinstance(cell.value, str):
            if "removal_photo" in cell.value:
                cell.value = ""
            if "manifest_photo" in cell.value:
                cell.value = ""

# 4. Save
if len(wb.sheetnames) > 1:
    wb.remove(source_sheet)

wb.save(OUTPUT_PATH)
print(f"Saved to {OUTPUT_PATH}")
print(f"Output Size: {os.path.getsize(OUTPUT_PATH)} bytes")
